from openpyxl import load_workbook
import pandas as pd
import imaplib
import imapclient
import pyzmail

assignment_substring = 'assignment'.lower()
deadline_substring = 'deadline'.lower()
the_substring = 'The'.lower()
on_substring = 'On'.lower()


def convertLower(line):
    new_line = line.lower()
    return new_line

def list_of_words(string_to_split):
    return string_to_split.split()

def is_deadline_assignment_present(email):
    
    'Checks if the words deadline and assignments are present in the email'
    
    if (assignment_substring in email.lower()) and (deadline_substring in email.lower()):
        return True
    
def assignment_titles(email):
    if is_deadline_assignment_present(email):
        for line in email.splitlines():
            new_line = convertLower(line)
            if assignment_substring in new_line:
                index_assignment = new_line.index(assignment_substring)
                list_of_word_on = list_of_words(new_line[index_assignment:])
                list_of_word = list_of_words(new_line[:(index_assignment+11)])
                
                'Ckecks every possible combination to accurately determine the title of the assignment'
                
                if the_substring in new_line[:index_assignment]:
                    return (list_of_word[list_of_word.index(the_substring)+1])
                elif len(list_of_word_on) > 3:
                    if (list_of_word_on[list_of_word_on.index(assignment_substring)+2]) == on_substring:
                        return (list_of_word_on[list_of_word_on.index(on_substring)+1])
                elif ':' not in (list_of_word[list_of_word.index(assignment_substring)-1]):
                    return (list_of_word[list_of_word.index(assignment_substring)-1])
                    break
                else:
                    return list_of_word_on[list_of_word_on.index(assignment_substring)+1]
                
def deadline_dates(email):
    if is_deadline_assignment_present(email):
        for line in email.splitlines():
            if deadline_substring in line:
                    index_deadline = line.index('deadline')
                    new_search = line[index_deadline:]
                    if 'is' in new_search:
                        index_is = line.index('is')
                        deadline = index_is+3
                        return line[deadline:]
                    else:
                        return line[deadline+8:]

def cell_exist(ws, ass_title, dead_date):
    to_return = True
    for row_num in range(1, ws.max_row):
        if ws.cell(row=row_num,column=1).value != ass_title and ws.cell(row=row_num,column=2).value != dead_date:
           to_return = True
        to_return = False
        
def main():
    passw = raw_input("Enter password: ")
    if passw:
        
        'Enter your email where you see YOUR EMAIL'
        'Connecting to the Imap server'
        
        imapObj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
        imapObj.login('Your email', passw)
        imapObj.select_folder('INBOX', readonly=True)
        UIDs = imapObj.search(['BODY', 'assignment'])
        rawMessages = imapObj.fetch(UIDs, ['BODY[]'])
        'Changing the mail to'
        for uuid in UIDs:
            email = ''
            assignment_title = ''
            deadline_date = ''
            message = pyzmail.PyzMessage.factory(rawMessages[uuid]['BODY[]'])
            
            'Change the email body to a text format'
            
            email = message.text_part.get_payload().decode(message.text_part.charset)
            assignment_title = assignment_titles(email)
            deadline_date = deadline_dates(email)
            new_row_data = [
            [assignment_title, deadline_date],
            ]
            
            'Make sure that you have this assignments.xlsx file on your pc'
            
            wb = load_workbook("assignments.xlsx")
            ws = wb.worksheets[0]
            for row_data in new_row_data:
                if cell_exist(ws, assignment_title, deadline_date) == True:
                    ws.append(row_data)
                    wb.save("assignments.xlsx")
                    

if __name__ == '__main__':
    main()

